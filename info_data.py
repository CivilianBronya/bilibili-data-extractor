import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

HEADERS = [
    "标题", "链接", "up主", "up主id", "精确播放数", "历史累计弹幕数", "点赞数", "投硬币枚数",
    "收藏人数", "转发人数", "发布时间", "视频时长(秒)", "视频简介", "作者简介", "标签", "视频aid", "爬取时间"
]

def init_excel(file_path):
    """初始化Excel文件，写入表头"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilibili视频信息"
    ws.append(HEADERS)
    wb.save(file_path)

def load_or_create_workbook(file_path):
    """打开已有Excel或新建"""
    if not os.path.exists(file_path):
        init_excel(file_path)
    return load_workbook(file_path)

def sanitize_value(value):
    """处理写入Excel的值，列表转字符串，None转空字符串"""
    if isinstance(value, list):
        return ",".join(str(v) for v in value)
    if value is None:
        return ""
    return str(value)

def to_int(value):
    """安全转换为整数，失败返回0"""
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

def append_video_info(file_path, video_info: dict):
    """
    写入视频信息到Excel，保留历史记录（不去重）
    :param file_path: Excel文件路径
    :param video_info: dict，视频信息字段应对应HEADERS
    """
    try:
        wb = load_or_create_workbook(file_path)
        ws = wb.active

        # 调试输出，确认传入数据
        print(f"[DEBUG] 准备写入 Excel: {video_info.get('url')}")
        print(f"[DEBUG] 标题: {video_info.get('title')}")
        print(f"[DEBUG] 播放量: {video_info.get('views')}, 点赞: {video_info.get('likes')}")

        values = [
            sanitize_value(video_info.get("title", "")),
            sanitize_value(video_info.get("url", "")),
            sanitize_value(video_info.get("author", "")),
            sanitize_value(video_info.get("author_id", "")),
            to_int(video_info.get("views")),
            to_int(video_info.get("danmaku")),
            to_int(video_info.get("likes")),
            to_int(video_info.get("coins")),
            to_int(video_info.get("favorites")),
            to_int(video_info.get("shares")),
            sanitize_value(video_info.get("publish_date", "")),
            to_int(video_info.get("duration")),
            sanitize_value(video_info.get("video_desc", "")),
            sanitize_value(video_info.get("author_desc", "")),
            sanitize_value(video_info.get("tags", "")),
            sanitize_value(video_info.get("video_aid", "")),
            datetime.now().isoformat(timespec='seconds'),
        ]

        ws.append(values)
        wb.save(file_path)
        print(f"[Excel] 已写入: {video_info.get('url')}")

        return True

    except Exception as e:
        print(f"[ERROR] 写入 Excel 失败: {e}")
        return False
